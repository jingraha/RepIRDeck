from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

from .charts import render_workbook_charts
from .config import (
    COLORS,
    DISPLAY_SLIDE_NUMBER,
    OUTPUT_DIR,
    XLSX_PATH,
)
from .model import ModelBundle
from .sources import Source, source_rows


SHEET_NAMES = [
    "Assumptions",
    "Visuals",
    "P&L",
    "Cash Flow",
    "Balance Sheet",
    "Read Me",
    "Source Register",
    "Public Facts",
    "Historical KPIs",
    "Revenue Build",
    "Cohorts & Retention",
    "Unit Economics",
    "GTM & Pipeline",
    "Market Model",
    "Scenarios",
    "Use of Proceeds",
    "Valuation",
    "Slide Reconciliation",
    "QA Checks",
]


ORANGE_FILL = PatternFill("solid", fgColor=COLORS["orange"])
DARK_FILL = PatternFill("solid", fgColor=COLORS["ink"])
CREAM_FILL = PatternFill("solid", fgColor=COLORS["cream"])
BLUE_FILL = PatternFill("solid", fgColor=COLORS["orange_pale"])
ESTIMATE_FILL = PatternFill("solid", fgColor="FFF0E7")
GRAY_FILL = PatternFill("solid", fgColor=COLORS["gray_100"])
GREEN_FILL = PatternFill("solid", fgColor="FFE9DF")
RED_FILL = PatternFill("solid", fgColor="FFD4C2")
WHITE_FONT = Font(name="Arial", size=10, bold=True, color=COLORS["white"])
HEADER_FONT = Font(name="Arial", size=10, bold=True, color=COLORS["ink"])
BODY_FONT = Font(name="Arial", size=9, color=COLORS["charcoal"])
TITLE_FONT = Font(name="Arial", size=20, bold=True, color=COLORS["ink"])
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color=COLORS["gray_700"])
THIN_BORDER = Border(bottom=Side(style="thin", color=COLORS["gray_200"]))


def _sheet_title(
    ws: Any,
    title: str,
    subtitle: str,
    *,
    start_row: int = 1,
) -> int:
    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).font = TITLE_FONT
    ws.cell(start_row + 1, 1, subtitle)
    ws.cell(start_row + 1, 1).font = SUBTITLE_FONT
    return start_row + 3


def _style_header(ws: Any, row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = ws.cell(row, column)
        cell.fill = DARK_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22


def _write_table(
    ws: Any,
    start_row: int,
    headers: list[str],
    rows: Iterable[Iterable[Any]],
    *,
    number_formats: dict[int, str] | None = None,
    estimate_rows: set[int] | None = None,
) -> int:
    for column, header in enumerate(headers, 1):
        ws.cell(start_row, column, header)
    _style_header(ws, start_row, len(headers))

    last_row = start_row
    for row_offset, values in enumerate(rows, 1):
        excel_row = start_row + row_offset
        last_row = excel_row
        for column, value in enumerate(values, 1):
            cell = ws.cell(excel_row, column, value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if number_formats and column in number_formats:
                cell.number_format = number_formats[column]
            if estimate_rows and excel_row in estimate_rows:
                cell.fill = ESTIMATE_FILL
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{last_row}"
    return last_row


def _set_widths(ws: Any, widths: dict[int, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width


def _prepare_sheet(ws: Any, freeze: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = freeze
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.outlinePr.summaryBelow = True


def _flatten(
    payload: Any,
    prefix: str = "",
) -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    if isinstance(payload, dict):
        for key, value in payload.items():
            path = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(_flatten(value, path))
    elif isinstance(payload, list):
        for position, value in enumerate(payload):
            path = f"{prefix}[{position}]"
            rows.extend(_flatten(value, path))
    else:
        rows.append((prefix, payload))
    return rows


def _quote_sheet(name: str) -> str:
    return f"'{name}'"


def _read_me_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws)
    row = _sheet_title(
        ws,
        "Replit Series E Sourcebook",
        "Public evidence, analyst assumptions, operating model, scenarios, and slide-level reconciliation.",
    )
    ws.cell(row, 1, "Important")
    ws.cell(row, 1).font = Font(name="Arial", size=12, bold=True, color=COLORS["orange"])
    ws.cell(
        row + 1,
        1,
        "This workbook is an illustrative financing model prepared from public information. "
        "It is not Replit confidential information, a forecast issued by Replit, or investment advice.",
    )
    ws.cell(row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 2, end_column=6)

    legend_row = row + 4
    _write_table(
        ws,
        legend_row,
        ["Evidence class", "Meaning", "Workbook color"],
        [
            ("Public fact", "Direct or attributable public disclosure", "Light orange"),
            ("Derived", "Arithmetic based on public facts", "Gray"),
            ("Est.", "Analyst estimate based on stated assumptions", "Orange"),
            ("Management target", "Forward-looking goal stated by management", "Dark orange"),
        ],
    )
    fills = [BLUE_FILL, GRAY_FILL, ESTIMATE_FILL, GREEN_FILL]
    for offset, fill in enumerate(fills, 1):
        for column in range(1, 4):
            ws.cell(legend_row + offset, column).fill = fill

    quick_row = legend_row + 7
    tx = model.transaction
    _write_table(
        ws,
        quick_row,
        ["Transaction metric", "Value", "Evidence"],
        [
            ("Round", tx["round"], "Illustrative"),
            ("Total raise", tx["total_raise"], "Illustrative"),
            ("Primary capital", tx["primary"], "Illustrative"),
            ("Secondary liquidity", tx["secondary"], "Illustrative"),
            ("Pre-money valuation", tx["pre_money"], "Illustrative"),
            ("Post-money valuation", tx["post_money"], "Derived"),
            ("Primary dilution", tx["primary_ownership"], "Derived"),
            ("Total buyer ownership", tx["total_buyer_ownership"], "Derived"),
        ],
        number_formats={2: '$#,##0.0;[Red]($#,##0.0)'},
    )
    for excel_row in range(quick_row + 1, quick_row + 9):
        ws.cell(excel_row, 2).fill = ESTIMATE_FILL
    ws.cell(quick_row + 7, 2).number_format = "0.0%"
    ws.cell(quick_row + 8, 2).number_format = "0.0%"
    _set_widths(ws, {1: 28, 2: 20, 3: 22, 4: 14, 5: 14, 6: 14})
    return {
        "transaction.total_raise": f"{_quote_sheet(ws.title)}!$B${quick_row + 2}",
        "transaction.primary": f"{_quote_sheet(ws.title)}!$B${quick_row + 3}",
        "transaction.secondary": f"{_quote_sheet(ws.title)}!$B${quick_row + 4}",
        "transaction.pre_money": f"{_quote_sheet(ws.title)}!$B${quick_row + 5}",
        "transaction.post_money": f"{_quote_sheet(ws.title)}!$B${quick_row + 6}",
        "transaction.primary_ownership": f"{_quote_sheet(ws.title)}!$B${quick_row + 7}",
        "transaction.total_buyer_ownership": f"{_quote_sheet(ws.title)}!$B${quick_row + 8}",
    }


def _source_register_sheet(ws: Any, sources: dict[str, Source]) -> None:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Source Register",
        "Canonical public and modeled sources. URLs are retained for diligence and slide reconciliation.",
    )
    rows = []
    for source in sources.values():
        rows.append(
            (
                source.id,
                source.publisher,
                source.title,
                source.publication_date,
                source.access_date,
                source.quality,
                source.evidence_class,
                source.url,
            )
        )
    _write_table(
        ws,
        start,
        [
            "Source ID",
            "Publisher",
            "Title",
            "Publication date",
            "Access date",
            "Quality",
            "Evidence class",
            "URL",
        ],
        rows,
    )
    for row in range(start + 1, start + 1 + len(rows)):
        url_cell = ws.cell(row, 8)
        if url_cell.value:
            url_cell.hyperlink = url_cell.value
            url_cell.style = "Hyperlink"
        evidence = ws.cell(row, 7).value
        fill = ESTIMATE_FILL if evidence == "estimate" else BLUE_FILL
        for column in range(1, 9):
            ws.cell(row, column).fill = fill
    _set_widths(ws, {1: 22, 2: 22, 3: 52, 4: 16, 5: 14, 6: 22, 7: 18, 8: 75})


def _public_facts_sheet(ws: Any, sources: dict[str, Source]) -> None:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Public Facts",
        "Claim-level evidence extracted from the source register. Modeled claims remain visibly separate.",
    )
    rows = source_rows(sources)
    _write_table(
        ws,
        start,
        [
            "Source ID",
            "Publisher",
            "Title",
            "Claim",
            "Publication date",
            "Access date",
            "Quality",
            "Evidence class",
            "URL",
        ],
        [
            (
                row["source_id"],
                row["publisher"],
                row["title"],
                row["claim"],
                row["publication_date"],
                row["access_date"],
                row["quality"],
                row["evidence_class"],
                row["url"],
            )
            for row in rows
        ],
    )
    for row in range(start + 1, start + 1 + len(rows)):
        fill = (
            ESTIMATE_FILL
            if ws.cell(row, 8).value == "estimate"
            else BLUE_FILL
        )
        for column in range(1, 10):
            ws.cell(row, column).fill = fill
        if ws.cell(row, 9).value:
            ws.cell(row, 9).hyperlink = ws.cell(row, 9).value
            ws.cell(row, 9).style = "Hyperlink"
    _set_widths(ws, {1: 22, 2: 22, 3: 48, 4: 70, 5: 16, 6: 14, 7: 22, 8: 18, 9: 70})


def _assumptions_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Assumptions",
        "Editable structured inputs. Orange rows are analyst assumptions. Public anchors are documented separately.",
    )
    flattened = _flatten(model.assumptions)
    rows: list[tuple[Any, ...]] = []
    for key, value in flattened:
        if isinstance(value, bool):
            kind = "Boolean"
        elif isinstance(value, (int, float)):
            kind = "Numeric"
        elif value is None:
            kind = "Blank"
        else:
            kind = "Text"
        rows.append((key, value, kind, "Est. unless identified in Source Register"))
    _write_table(
        ws,
        start,
        ["Assumption key", "Value", "Type", "Evidence note"],
        rows,
        estimate_rows=set(range(start + 1, start + 1 + len(rows))),
    )
    lookup: dict[str, str] = {}
    for offset, (key, _) in enumerate(flattened, 1):
        lookup[key] = f"{_quote_sheet(ws.title)}!$B${start + offset}"
    _set_widths(ws, {1: 58, 2: 22, 3: 14, 4: 44})
    return lookup


def _historical_kpis_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Historical KPIs",
        "Public anchors and analyst estimates from 2023 to 2026.",
    )
    metrics = [
        ("Revenue", "revenue", "$M", "$#,##0"),
        ("Exit ARR", "exit_arr", "$M", "$#,##0"),
        ("Registered users", "registered_users_m", "M", "0.0"),
        ("Monthly active builders", "monthly_active_builders_m", "M", "0.0"),
        ("Paid individuals", "paid_individuals_m", "M", "0.00"),
        ("Enterprise customers", "enterprise_customers", "#", "#,##0"),
        ("Paid conversion", "paid_conversion", "%", "0.0%"),
        ("Enterprise GRR", "enterprise_grr", "%", "0.0%"),
        ("Enterprise NRR", "enterprise_nrr", "%", "0.0%"),
        ("International revenue share", "international_revenue_share", "%", "0.0%"),
    ]
    years = [2023, 2024, 2025, 2026]
    headers = ["Metric", "Units", *years, "Evidence"]
    rows = []
    for label, column, units, _ in metrics:
        values = [model.financials.loc[year, column] for year in years]
        rows.append((label, units, *values, "Public anchor / Est."))
    _write_table(ws, start, headers, rows)
    lookup: dict[str, str] = {}
    for row_offset, (_, column, _, number_format) in enumerate(metrics, 1):
        for year_offset, year in enumerate(years, 3):
            cell = ws.cell(start + row_offset, year_offset)
            cell.number_format = number_format
            cell.fill = ESTIMATE_FILL
            lookup[f"kpi.{column}.{year}"] = (
                f"{_quote_sheet(ws.title)}!${get_column_letter(year_offset)}${start + row_offset}"
            )
    _set_widths(ws, {1: 32, 2: 10, 3: 14, 4: 14, 5: 14, 6: 14, 7: 24})
    return lookup


def _revenue_build_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Revenue Build",
        "Segment model in $M. The 2023 to 2026 periods are estimated history. The forecast starts in 2027.",
    )
    years = list(model.financials.index)
    headers = ["Metric", *years]
    metric_rows = [
        ("Individual", "revenue_individual", "$#,##0"),
        ("Enterprise", "revenue_enterprise", "$#,##0"),
        ("Usage & deploy", "revenue_usage_and_deploy", "$#,##0"),
        ("Total revenue", "revenue", "$#,##0"),
        ("Year-over-year growth", "revenue_growth", "0.0%"),
    ]
    rows = []
    for label, column, _ in metric_rows:
        rows.append((label, *model.financials[column].tolist()))
    _write_table(ws, start, headers, rows)
    lookup: dict[str, str] = {}
    for row_offset, (_, column, number_format) in enumerate(metric_rows, 1):
        for year_offset, year in enumerate(years, 2):
            cell = ws.cell(start + row_offset, year_offset)
            cell.number_format = number_format
            cell.fill = ESTIMATE_FILL
            lookup[f"financial.{column}.{year}"] = (
                f"{_quote_sheet(ws.title)}!${get_column_letter(year_offset)}${start + row_offset}"
            )
    chart = LineChart()
    chart.title = "Revenue by segment"
    chart.y_axis.title = "$M"
    chart.x_axis.title = "Year"
    chart.style = 13
    chart.height = 8
    chart.width = 18
    data = Reference(
        ws,
        min_col=2,
        max_col=1 + len(years),
        min_row=start + 1,
        max_row=start + 3,
    )
    categories = Reference(
        ws,
        min_col=2,
        max_col=1 + len(years),
        min_row=start,
        max_row=start,
    )
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(categories)
    for series, color in zip(
        chart.series,
        ("orange", "orange_mid", "coral"),
        strict=True,
    ):
        series.graphicalProperties.line.solidFill = COLORS[color]
        series.graphicalProperties.line.width = 22000
    ws.add_chart(chart, f"A{start + 8}")
    _set_widths(ws, {1: 30, **{column: 14 for column in range(2, 11)}})
    return lookup


def _cohorts_sheet(ws: Any, model: ModelBundle) -> None:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Cohorts & Retention",
        "Illustrative enterprise ARR cohorts. All values are analyst estimates.",
    )
    pivot = model.cohorts.pivot(index="cohort", columns="month", values="multiple")
    rows = []
    for cohort, values in pivot.iterrows():
        rows.append(
            (
                cohort,
                1.0,
                values.get(12),
                values.get(24),
                values.get(36),
                "Est.",
            )
        )
    _write_table(
        ws,
        start,
        ["Cohort", "Start", "Month 12", "Month 24", "Month 36", "Evidence"],
        rows,
    )
    for row in range(start + 1, start + 1 + len(rows)):
        for column in range(2, 6):
            ws.cell(row, column).number_format = "0.00x"
            ws.cell(row, column).fill = ESTIMATE_FILL
    _set_widths(ws, {1: 14, 2: 14, 3: 14, 4: 14, 5: 14, 6: 18})


def _unit_economics_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Unit Economics",
        "Analyst estimates connecting model efficiency, gross margin, acquisition cost, and payback.",
    )
    years = list(model.unit_economics.index)
    metrics = [
        ("Blended gross margin", "blended_gross_margin", "0.0%"),
        ("Individual CAC", "individual_cac_usd", '$0'),
        ("Individual payback", "individual_payback_months", '0.0 "mo"'),
        ("Enterprise CAC", "enterprise_cac_k", '$0 "k"'),
        ("Enterprise payback", "enterprise_payback_months", '0 "mo"'),
        ("Enterprise LTV / CAC", "enterprise_ltv_cac", "0.0x"),
        (
            "Compute cost / successful build index",
            "compute_cost_per_successful_build_index",
            '0 "index"',
        ),
    ]
    rows = [
        (label, *model.unit_economics[column].tolist(), "Est.")
        for label, column, _ in metrics
    ]
    _write_table(ws, start, ["Metric", *years, "Evidence"], rows)
    lookup: dict[str, str] = {}
    for row_offset, (_, column, number_format) in enumerate(metrics, 1):
        for year_offset, year in enumerate(years, 2):
            cell = ws.cell(start + row_offset, year_offset)
            if column == "net_income":
                operating_income_row = start + 9
                cell.value = f"={get_column_letter(year_offset)}{operating_income_row}"
            cell.number_format = number_format
            cell.fill = ESTIMATE_FILL
            lookup[f"unit.{column}.{year}"] = (
                f"{_quote_sheet(ws.title)}!${get_column_letter(year_offset)}${start + row_offset}"
            )
    _set_widths(ws, {1: 42, **{column: 14 for column in range(2, 7)}, 7: 16})
    return lookup


def _pipeline_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "GTM & Pipeline",
        "Illustrative Q3 2026 enterprise pipeline and sales-capacity assumptions.",
    )
    pipeline = model.assumptions["pipeline"]
    metrics = [
        ("Qualified pipeline", "qualified_pipeline", "$M", "$#,##0"),
        ("Late-stage pipeline", "late_stage_pipeline", "$M", "$#,##0"),
        ("Weighted pipeline", "weighted_pipeline", "$M", "$#,##0"),
        ("Next-12-month new ARR target", "next_12m_new_arr_target", "$M", "$#,##0"),
        ("Pipeline coverage", "coverage", "x", "0.0x"),
        ("Win rate", "win_rate", "%", "0.0%"),
        ("Median sales cycle", "median_sales_cycle_days", "days", '0 "days"'),
        ("Quota attainment", "quota_attainment", "%", "0.0%"),
        ("Fully ramped reps", "fully_ramped_reps", "#", "#,##0"),
        ("New ARR per ramped rep", "new_arr_per_ramped_rep", "$M", '$0.0 "M"'),
        ("Rep ramp", "rep_ramp_months", "months", '0 "mo"'),
    ]
    rows = [
        (label, pipeline[key], units, "Est.")
        for label, key, units, _ in metrics
    ]
    _write_table(ws, start, ["Metric", "Value", "Units", "Evidence"], rows)
    lookup: dict[str, str] = {}
    for row_offset, (_, key, _, number_format) in enumerate(metrics, 1):
        ws.cell(start + row_offset, 2).number_format = number_format
        ws.cell(start + row_offset, 2).fill = ESTIMATE_FILL
        lookup[f"pipeline.{key}"] = (
            f"{_quote_sheet(ws.title)}!$B${start + row_offset}"
        )
    _set_widths(ws, {1: 38, 2: 18, 3: 16, 4: 16})
    return lookup


def _market_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Market Model",
        "Bottom-up opportunity sizing. Population and monetization inputs are analyst estimates triangulated to public research.",
    )
    rows = [
        (
            row["name"],
            row["population_m"],
            row["annual_revenue_per_unit"],
            row["tam"],
            row["sam"],
            "Derived / Est.",
        )
        for _, row in model.market.iterrows()
    ]
    total_tam = model.market["tam"].sum()
    total_sam = model.market["sam"].sum()
    rows.append(("Total", None, None, total_tam, total_sam, "Derived"))
    _write_table(
        ws,
        start,
        [
            "Segment",
            "Population / workloads (M)",
            "Annual revenue per unit ($)",
            "TAM ($M)",
            "Near-term SAM ($M)",
            "Evidence",
        ],
        rows,
    )
    for row in range(start + 1, start + 1 + len(rows)):
        ws.cell(row, 2).number_format = "0.0"
        ws.cell(row, 3).number_format = "$#,##0"
        ws.cell(row, 4).number_format = "$#,##0"
        ws.cell(row, 5).number_format = "$#,##0"
        fill = GRAY_FILL if row == start + len(rows) else ESTIMATE_FILL
        for column in range(1, 7):
            ws.cell(row, column).fill = fill
    total_row = start + len(rows)
    _set_widths(ws, {1: 34, 2: 24, 3: 26, 4: 18, 5: 22, 6: 18})
    return {
        "market.total_tam": f"{_quote_sheet(ws.title)}!$D${total_row}",
        "market.total_sam": f"{_quote_sheet(ws.title)}!$E${total_row}",
    }


def _p_and_l_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "P&L",
        "Income statement in $M. All periods are analyst estimates.",
    )
    years = list(model.income_statement.index)
    metrics = [
        ("Revenue", "revenue", "$#,##0"),
        ("Growth", "revenue_growth", "0.0%"),
        ("Cost of revenue", "cost_of_revenue", "$#,##0"),
        ("Gross profit", "gross_profit", "$#,##0"),
        ("Gross margin", "gross_margin", "0.0%"),
        ("R&D", "r_and_d", "$#,##0"),
        ("Sales & marketing", "sales_and_marketing", "$#,##0"),
        ("G&A", "g_and_a", "$#,##0"),
        ("Operating income", "adjusted_operating_income", "$#,##0;[Red]($#,##0)"),
        ("Operating margin", "adjusted_operating_margin", "0.0%;[Red](0.0%)"),
        ("Net income", "net_income", "$#,##0;[Red]($#,##0)"),
        ("Stock compensation included in expenses", "stock_compensation", "$#,##0"),
    ]
    rows = [
        (label, *model.income_statement[column].tolist())
        for label, column, _ in metrics
    ]
    _write_table(ws, start, ["Metric", *years], rows)
    lookup: dict[str, str] = {}
    for row_offset, (_, column, number_format) in enumerate(metrics, 1):
        for year_offset, year in enumerate(years, 2):
            cell = ws.cell(start + row_offset, year_offset)
            cell.number_format = number_format
            cell.fill = ESTIMATE_FILL
            lookup[f"financial.{column}.{year}"] = (
                f"{_quote_sheet(ws.title)}!${get_column_letter(year_offset)}${start + row_offset}"
            )
    for row_offset in (1, 4, 9, 11):
        for column in range(1, len(years) + 2):
            ws.cell(start + row_offset, column).font = Font(
                name="Arial",
                size=9,
                bold=True,
                color=COLORS["ink"],
            )
    _set_widths(ws, {1: 34, **{column: 14 for column in range(2, 11)}})
    return lookup


def _cash_flow_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Cash Flow",
        "Annual cash flow and monthly cash runway in $M.",
    )
    annual_metrics = [
        ("Net income", "net_income", "$#,##0;[Red]($#,##0)"),
        ("Stock compensation", "stock_compensation", "$#,##0"),
        ("Change in working capital", "change_in_working_capital", "$#,##0;[Red]($#,##0)"),
        ("Cash from operations", "cash_from_operations", "$#,##0;[Red]($#,##0)"),
        ("Capital expenditures", "capital_expenditures", "$#,##0;[Red]($#,##0)"),
        ("Free cash flow", "free_cash_flow", "$#,##0;[Red]($#,##0)"),
        ("Prior equity financing", "prior_equity_financing", "$#,##0"),
        ("Series E primary", "series_e_primary", "$#,##0"),
        ("Financing fees", "financing_fees", "$#,##0;[Red]($#,##0)"),
        ("Net change in cash", "net_change_in_cash", "$#,##0;[Red]($#,##0)"),
        ("Beginning cash", "beginning_cash", "$#,##0"),
        ("Ending cash", "ending_cash", "$#,##0"),
    ]
    years = list(model.cash_flow_statement.index)
    annual_rows = [
        (label, *model.cash_flow_statement[column].tolist())
        for label, column, _ in annual_metrics
    ]
    _write_table(ws, start, ["Annual cash flow", *years], annual_rows)
    lookup: dict[str, str] = {}
    row_by_metric = {
        column: start + row_offset
        for row_offset, (_, column, _) in enumerate(annual_metrics, 1)
    }
    pnl_row_by_metric = {
        "net_income": 15,
        "stock_compensation": 16,
    }
    for row_offset, (_, column, number_format) in enumerate(annual_metrics, 1):
        for year_offset, year in enumerate(years, 2):
            cell = ws.cell(start + row_offset, year_offset)
            column_letter = get_column_letter(year_offset)
            if column in pnl_row_by_metric:
                cell.value = (
                    f"='P&L'!{column_letter}{pnl_row_by_metric[column]}"
                )
            elif column == "cash_from_operations":
                cell.value = (
                    f"={column_letter}{row_by_metric['net_income']}"
                    f"+{column_letter}{row_by_metric['stock_compensation']}"
                    f"+{column_letter}{row_by_metric['change_in_working_capital']}"
                )
            elif column == "free_cash_flow":
                cell.value = (
                    f"={column_letter}{row_by_metric['cash_from_operations']}"
                    f"+{column_letter}{row_by_metric['capital_expenditures']}"
                )
            elif column == "net_change_in_cash":
                cell.value = (
                    f"={column_letter}{row_by_metric['free_cash_flow']}"
                    f"+{column_letter}{row_by_metric['prior_equity_financing']}"
                    f"+{column_letter}{row_by_metric['series_e_primary']}"
                    f"+{column_letter}{row_by_metric['financing_fees']}"
                )
            elif column == "beginning_cash" and year_offset > 2:
                prior_column = get_column_letter(year_offset - 1)
                cell.value = f"={prior_column}{row_by_metric['ending_cash']}"
            elif column == "ending_cash":
                cell.value = (
                    f"={column_letter}{row_by_metric['beginning_cash']}"
                    f"+{column_letter}{row_by_metric['net_change_in_cash']}"
                )
            cell.number_format = number_format
            cell.fill = ESTIMATE_FILL
            lookup[f"cash_flow.{column}.{year}"] = (
                f"{_quote_sheet(ws.title)}!${get_column_letter(year_offset)}${start + row_offset}"
            )

    summary = model.cash_summary
    summary_start = start + len(annual_metrics) + 3
    summary_rows = [
        ("Cash at September 2026", summary["beginning_cash"], "Est."),
        ("Minimum operating cash", summary["minimum_cash"], "Est."),
        ("Series E close", summary["financing_close_date"].to_pydatetime(), "Est."),
        ("Net primary proceeds", summary["net_primary_proceeds"], "Derived"),
        ("Cash falls below minimum without Series E", summary["funding_need_date"].to_pydatetime(), "Derived"),
        ("IPO target", summary["ipo_target_date"].to_pydatetime(), "Est."),
        ("Cash at IPO in base case", summary["cash_at_ipo"], "Derived"),
        ("Cash at IPO in downside case", summary["downside_cash_at_ipo"], "Derived"),
        ("Next equity need", summary["next_equity_need"], "Derived"),
    ]
    _write_table(
        ws,
        summary_start,
        ["Cash summary", "Value", "Evidence"],
        summary_rows,
    )
    for row in range(summary_start + 1, summary_start + 1 + len(summary_rows)):
        ws.cell(row, 2).fill = ESTIMATE_FILL
    ws.cell(summary_start + 1, 2).number_format = "$#,##0"
    ws.cell(summary_start + 2, 2).number_format = "$#,##0"
    ws.cell(summary_start + 3, 2).number_format = "mmm-yy"
    ws.cell(summary_start + 4, 2).number_format = "$#,##0"
    ws.cell(summary_start + 5, 2).number_format = "mmm-yy"
    ws.cell(summary_start + 6, 2).number_format = "mmm-yy"
    ws.cell(summary_start + 7, 2).number_format = "$#,##0"
    ws.cell(summary_start + 8, 2).number_format = "$#,##0"

    table_start = summary_start + len(summary_rows) + 3
    columns = [
        ("Month", None),
        ("Beginning cash", "$#,##0"),
        ("Revenue", "$#,##0"),
        ("Gross profit", "$#,##0"),
        ("Operating expenses", "$#,##0"),
        ("Adjusted operating income", "$#,##0;[Red]($#,##0)"),
        ("Stock compensation", "$#,##0"),
        ("Capital expenditures", "$#,##0"),
        ("Working capital change", "$#,##0"),
        ("Free cash flow", "$#,##0;[Red]($#,##0)"),
        ("Series E primary", "$#,##0"),
        ("Financing fees", "$#,##0"),
        ("Ending cash", "$#,##0"),
        ("Ending cash without Series E", "$#,##0;[Red]($#,##0)"),
        ("Downside ending cash", "$#,##0;[Red]($#,##0)"),
        ("Minimum cash", "$#,##0"),
        ("Milestone", None),
    ]
    rows = []
    for date, row in model.monthly_cash_flow.iterrows():
        rows.append(
            (
                date.to_pydatetime(),
                row["beginning_cash"],
                row["revenue"],
                row["gross_profit"],
                row["operating_expenses"],
                row["adjusted_operating_income"],
                row["stock_compensation"],
                row["capital_expenditures"],
                row["working_capital_change"],
                row["free_cash_flow"],
                row["primary_financing"],
                row["financing_fees"],
                row["ending_cash"],
                row["ending_cash_without_series_e"],
                row["ending_cash_downside"],
                row["minimum_cash"],
                row["milestone"],
            )
        )
    _write_table(
        ws,
        table_start,
        [column[0] for column in columns],
        rows,
    )
    for row_offset in range(1, len(rows) + 1):
        excel_row = table_start + row_offset
        ws.cell(excel_row, 1).number_format = "mmm-yy"
        for column_index, (_, number_format) in enumerate(columns[1:], 2):
            if number_format:
                ws.cell(excel_row, column_index).number_format = number_format
            ws.cell(excel_row, column_index).fill = ESTIMATE_FILL

    chart = LineChart()
    chart.title = "Monthly ending cash"
    chart.y_axis.title = "$M"
    chart.height = 9
    chart.width = 18
    chart.legend.position = "b"
    chart.x_axis.tickLblSkip = 6
    chart.x_axis.number_format = "mmm-yy"
    chart.x_axis.tickLblPos = "low"
    data = Reference(
        ws,
        min_col=13,
        max_col=16,
        min_row=table_start,
        max_row=table_start + len(rows),
    )
    categories = Reference(
        ws,
        min_col=1,
        min_row=table_start + 1,
        max_row=table_start + len(rows),
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    for series, color in zip(
        chart.series,
        ("orange", "gray_500", "coral", "orange_dark"),
        strict=True,
    ):
        series.graphicalProperties.line.solidFill = COLORS[color]
        series.graphicalProperties.line.width = 22000
    ws.add_chart(chart, "J4")
    _set_widths(
        ws,
        {
            1: 14,
            2: 18,
            3: 14,
            4: 16,
            5: 20,
            6: 24,
            7: 20,
            8: 22,
            9: 22,
            10: 18,
            11: 18,
            12: 16,
            13: 18,
            14: 30,
            15: 22,
            16: 18,
            17: 34,
        },
    )
    return {
        **lookup,
        "cash.beginning_cash": f"{_quote_sheet(ws.title)}!$B${summary_start + 1}",
        "cash.minimum_cash": f"{_quote_sheet(ws.title)}!$B${summary_start + 2}",
        "cash.financing_close_date": f"{_quote_sheet(ws.title)}!$B${summary_start + 3}",
        "cash.net_primary_proceeds": f"{_quote_sheet(ws.title)}!$B${summary_start + 4}",
        "cash.funding_need_date": f"{_quote_sheet(ws.title)}!$B${summary_start + 5}",
        "cash.ipo_target_date": f"{_quote_sheet(ws.title)}!$B${summary_start + 6}",
        "cash.cash_at_ipo": f"{_quote_sheet(ws.title)}!$B${summary_start + 7}",
        "cash.downside_cash_at_ipo": f"{_quote_sheet(ws.title)}!$B${summary_start + 8}",
        "cash.next_equity_need": f"{_quote_sheet(ws.title)}!$B${summary_start + 9}",
    }


def _balance_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Balance Sheet",
        "Simplified balance sheet in $M. Cash links to the cash-flow statement.",
    )
    years = list(model.balance_sheet.index)
    metrics = [
        ("Cash", "cash"),
        ("Accounts receivable", "accounts_receivable"),
        ("Other current assets", "other_current_assets"),
        ("Property and equipment", "property_and_equipment"),
        ("Other assets", "other_assets"),
        ("Total assets", "total_assets"),
        ("Accounts payable", "accounts_payable"),
        ("Accrued expenses", "accrued_expenses"),
        ("Deferred revenue", "deferred_revenue"),
        ("Other liabilities", "other_liabilities"),
        ("Debt", "total_debt"),
        ("Total liabilities", "total_liabilities"),
        ("Total equity", "total_equity"),
        ("Liabilities and equity", "liabilities_and_equity"),
        ("Balance check", "balance_check"),
    ]
    rows = [
        (label, *model.balance_sheet[column].tolist())
        for label, column in metrics
    ]
    _write_table(ws, start, ["Balance sheet", *years], rows)
    lookup: dict[str, str] = {}
    row_by_metric = {
        column: start + row_offset
        for row_offset, (_, column) in enumerate(metrics, 1)
    }
    total_rows = {"total_assets", "total_liabilities", "total_equity", "liabilities_and_equity", "balance_check"}
    for row_offset, (_, column) in enumerate(metrics, 1):
        for year_offset, year in enumerate(years, 2):
            cell = ws.cell(start + row_offset, year_offset)
            column_letter = get_column_letter(year_offset)
            if column == "cash":
                cash_flow_column = get_column_letter(year_offset)
                cell.value = f"='Cash Flow'!{cash_flow_column}16"
            elif column == "total_assets":
                asset_rows = [
                    row_by_metric["cash"],
                    row_by_metric["accounts_receivable"],
                    row_by_metric["other_current_assets"],
                    row_by_metric["property_and_equipment"],
                    row_by_metric["other_assets"],
                ]
                cell.value = "=" + "+".join(
                    f"{column_letter}{row}" for row in asset_rows
                )
            elif column == "total_liabilities":
                liability_rows = [
                    row_by_metric["accounts_payable"],
                    row_by_metric["accrued_expenses"],
                    row_by_metric["deferred_revenue"],
                    row_by_metric["other_liabilities"],
                    row_by_metric["total_debt"],
                ]
                cell.value = "=" + "+".join(
                    f"{column_letter}{row}" for row in liability_rows
                )
            elif column == "total_equity":
                cell.value = (
                    f"={column_letter}{row_by_metric['total_assets']}"
                    f"-{column_letter}{row_by_metric['total_liabilities']}"
                )
            elif column == "liabilities_and_equity":
                cell.value = (
                    f"={column_letter}{row_by_metric['total_liabilities']}"
                    f"+{column_letter}{row_by_metric['total_equity']}"
                )
            elif column == "balance_check":
                cell.value = (
                    f"={column_letter}{row_by_metric['total_assets']}"
                    f"-{column_letter}{row_by_metric['liabilities_and_equity']}"
                )
            cell.number_format = "$#,##0;[Red]($#,##0)"
            cell.fill = ESTIMATE_FILL
            lookup[f"balance_sheet.{column}.{year}"] = (
                f"{_quote_sheet(ws.title)}!${get_column_letter(year_offset)}${start + row_offset}"
            )
        if column in total_rows:
            for column_index in range(1, len(years) + 2):
                ws.cell(start + row_offset, column_index).font = Font(
                    name="Arial",
                    size=9,
                    bold=True,
                    color=COLORS["ink"],
                )
    _set_widths(ws, {1: 32, **{column: 14 for column in range(2, 9)}})
    return lookup


def _visuals_sheet(ws: Any, model: ModelBundle) -> None:
    _prepare_sheet(ws)
    _sheet_title(
        ws,
        "Visuals",
        "Charts use the same data as the P&L, cash flow, balance sheet, and operating schedules.",
    )
    chart_paths = render_workbook_charts(model, OUTPUT_DIR / ".charts")
    revenue_image = XLImage(chart_paths["revenue"])
    revenue_image.width = 675
    revenue_image.height = 340
    ws.add_image(revenue_image, "A4")
    use_image = XLImage(chart_paths["use_of_proceeds"])
    use_image.width = 675
    use_image.height = 285
    ws.add_image(use_image, "J4")
    cash_image = XLImage(chart_paths["cash"])
    cash_image.width = 1360
    cash_image.height = 480
    ws.add_image(cash_image, "A23")

    revenue_start = 42
    years = list(range(2026, 2032))
    revenue_rows = [
        (
            f"{year}E",
            model.financials.loc[year, "revenue_individual"],
            model.financials.loc[year, "revenue_enterprise"],
            model.financials.loc[year, "revenue_usage_and_deploy"],
        )
        for year in years
    ]
    _write_table(
        ws,
        revenue_start,
        ["Year", "Individual", "Enterprise", "Agent use and hosting"],
        revenue_rows,
    )
    use_start = 42
    use_columns = ["Use", *model.use_of_proceeds["category"].tolist()]
    use_values = ["Primary", *model.use_of_proceeds["amount"].tolist()]
    for column, value in enumerate(use_columns, 10):
        ws.cell(use_start, column, value)
    for column, value in enumerate(use_values, 10):
        ws.cell(use_start + 1, column, value)
    cash_start = revenue_start + len(revenue_rows) + 4
    cash_rows = []
    for date, row in model.monthly_cash_flow.loc[
        : model.cash_summary["ipo_target_date"]
    ].iterrows():
        no_financing = (
            row["ending_cash_without_series_e"]
            if date <= model.cash_summary["funding_need_date"]
            else None
        )
        cash_rows.append(
            (
                date.strftime("%b-%y"),
                row["ending_cash"],
                row["ending_cash_downside"],
                no_financing,
                row["minimum_cash"],
            )
        )
    _write_table(
        ws,
        cash_start,
        [
            "Month",
            "Base with Series E",
            "Downside with Series E",
            "No Series E",
            "Minimum cash",
        ],
        cash_rows,
    )
    for row in range(cash_start + 1, cash_start + 1 + len(cash_rows)):
        for column in range(2, 17):
            ws.cell(row, column).number_format = "$#,##0"
    _set_widths(ws, {column: 18 for column in range(1, 17)})


def _scenarios_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Scenarios",
        "Base, downside, and upside operating outcomes from 2026 to 2031.",
    )
    years = list(model.scenarios["base"].index)
    rows: list[tuple[Any, ...]] = []
    order = ["downside", "base", "upside"]
    metrics = [
        ("Revenue", "revenue", "$#,##0"),
        ("Gross margin", "gross_margin", "0.0%"),
        ("Adjusted operating margin", "adjusted_operating_margin", "0.0%"),
        ("Free cash flow", "free_cash_flow", "$#,##0;[Red]($#,##0)"),
    ]
    for scenario_name in order:
        frame = model.scenarios[scenario_name]
        for label, column, _ in metrics:
            rows.append(
                (
                    scenario_name.title(),
                    label,
                    *frame[column].tolist(),
                )
            )
    _write_table(ws, start, ["Scenario", "Metric", *years], rows)
    lookup: dict[str, str] = {}
    for scenario_index, scenario_name in enumerate(order):
        for metric_index, (_, column, number_format) in enumerate(metrics):
            excel_row = start + 1 + scenario_index * len(metrics) + metric_index
            fill = {
                "downside": RED_FILL,
                "base": GRAY_FILL,
                "upside": GREEN_FILL,
            }[scenario_name]
            for year_offset, year in enumerate(years, 3):
                ws.cell(excel_row, year_offset).number_format = number_format
                ws.cell(excel_row, year_offset).fill = fill
                lookup[f"scenario.{scenario_name}.{column}.{year}"] = (
                    f"{_quote_sheet(ws.title)}!${get_column_letter(year_offset)}${excel_row}"
                )
    chart = LineChart()
    chart.title = "Revenue scenarios"
    chart.y_axis.title = "$M"
    chart.height = 8
    chart.width = 18
    revenue_rows = [start + 1, start + 5, start + 9]
    for row in revenue_rows:
        data = Reference(
            ws,
            min_col=3,
            max_col=2 + len(years),
            min_row=row,
            max_row=row,
        )
        chart.add_data(data, from_rows=True)
    categories = Reference(
        ws,
        min_col=3,
        max_col=2 + len(years),
        min_row=start,
        max_row=start,
    )
    chart.set_categories(categories)
    for series, color in zip(
        chart.series,
        ("orange_dark", "orange", "coral"),
        strict=True,
    ):
        series.graphicalProperties.line.solidFill = COLORS[color]
        series.graphicalProperties.line.width = 22000
    ws.add_chart(chart, f"A{start + 16}")
    _set_widths(ws, {1: 16, 2: 32, **{column: 14 for column in range(3, 9)}})
    return lookup


def _use_of_proceeds_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Use of Proceeds",
        "Allocation of $800M primary capital. The $200M secondary component is excluded.",
    )
    rows = [
        (
            row["category"],
            row["amount"],
            row["percent"],
            row["milestone"],
            "Est.",
        )
        for _, row in model.use_of_proceeds.iterrows()
    ]
    rows.append(
        (
            "Total primary",
            model.use_of_proceeds["amount"].sum(),
            model.use_of_proceeds["percent"].sum(),
            "Fully allocated",
            "Derived",
        )
    )
    _write_table(
        ws,
        start,
        ["Category", "Amount ($M)", "% of primary", "Milestone", "Evidence"],
        rows,
    )
    lookup: dict[str, str] = {}
    for row_offset, row in enumerate(rows, 1):
        ws.cell(start + row_offset, 2).number_format = "$#,##0"
        ws.cell(start + row_offset, 3).number_format = "0.0%"
        fill = GRAY_FILL if row_offset == len(rows) else ESTIMATE_FILL
        for column in range(1, 6):
            ws.cell(start + row_offset, column).fill = fill
        key = str(row[0]).lower().replace(" ", "_")
        lookup[f"use.{key}"] = (
            f"{_quote_sheet(ws.title)}!$B${start + row_offset}"
        )
    _set_widths(ws, {1: 30, 2: 18, 3: 18, 4: 62, 5: 16})
    return lookup


def _valuation_sheet(
    ws: Any,
    model: ModelBundle,
) -> dict[str, str]:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Valuation",
        "Illustrative forward-revenue framework. This is not a fairness opinion or investment recommendation.",
    )
    transaction = model.transaction
    ntm_year = model.assumptions["valuation"]["ntm_revenue_year"]
    ntm_revenue = model.financials.loc[ntm_year, "revenue"]
    metrics = [
        ("Pre-money valuation", transaction["pre_money"], "$M"),
        ("Post-money valuation", transaction["post_money"], "$M"),
        (f"{ntm_year} revenue", ntm_revenue, "$M"),
        ("Pre-money / NTM revenue", transaction["pre_money"] / ntm_revenue, "x"),
        ("Post-money / NTM revenue", transaction["post_money"] / ntm_revenue, "x"),
    ]
    _write_table(
        ws,
        start,
        ["Entry metric", "Value", "Units"],
        metrics,
    )
    for row in range(start + 1, start + 1 + len(metrics)):
        ws.cell(row, 2).fill = ESTIMATE_FILL
        ws.cell(row, 2).number_format = (
            "0.0x" if ws.cell(row, 3).value == "x" else "$#,##0"
        )

    sensitivity_start = start + 8
    exit_year = model.assumptions["valuation"]["exit_year"]
    exit_revenue = model.financials.loc[exit_year, "revenue"]
    multiples = model.assumptions["valuation"]["exit_revenue_multiples"]
    sensitivity_rows = []
    for multiple in multiples:
        equity_value = exit_revenue * multiple
        gross_moic = equity_value / transaction["post_money"]
        sensitivity_rows.append((multiple, exit_revenue, equity_value, gross_moic))
    _write_table(
        ws,
        sensitivity_start,
        ["Exit revenue multiple", f"{exit_year} revenue", "Implied equity value", "Gross MOIC"],
        sensitivity_rows,
    )
    for row in range(sensitivity_start + 1, sensitivity_start + 1 + len(sensitivity_rows)):
        ws.cell(row, 1).number_format = "0.0x"
        ws.cell(row, 2).number_format = "$#,##0"
        ws.cell(row, 3).number_format = "$#,##0"
        ws.cell(row, 4).number_format = "0.0x"
        for column in range(1, 5):
            ws.cell(row, column).fill = ESTIMATE_FILL

    comps_start = sensitivity_start + 8
    comps = [
        ("Salesforce", 4.5, "Forward P/S", "COMP-CRM-2026", "Public market"),
        ("Atlassian", 7.5, "Market cap / LTM revenue", "COMP-TEAM-2026", "Public proxy"),
        ("GitLab", 7.7, "Market cap / LTM revenue", "COMP-GTLB-2026", "Public proxy"),
        ("ServiceNow", 8.5, "Forward P/S", "COMP-NOW-2026", "Public market"),
        ("Microsoft", 11.4, "Market cap / LTM revenue", "COMP-MSFT-2026", "Public proxy"),
        (
            "Replit Series E",
            transaction["pre_money"] / ntm_revenue,
            "Pre-money / 2027E revenue",
            "MODEL-ESTIMATES",
            "Illustrative entry",
        ),
        ("Cursor", 29.3, "Valuation / annualized revenue", "CURSOR-D-2025", "Private-company disclosure"),
        ("Cloudflare", 40.4, "Market cap / LTM revenue", "COMP-NET-2026", "High-growth public proxy"),
    ]
    _write_table(
        ws,
        comps_start,
        ["Comparable", "Revenue multiple", "Basis", "Source ID", "Caveat"],
        comps,
    )
    for row in range(comps_start + 1, comps_start + 1 + len(comps)):
        ws.cell(row, 2).number_format = "0.0x"
        fill = ESTIMATE_FILL if ws.cell(row, 1).value == "Replit Series E" else BLUE_FILL
        for column in range(1, 6):
            ws.cell(row, column).fill = fill
    _set_widths(ws, {1: 34, 2: 20, 3: 20, 4: 18})
    return {
        "valuation.entry_multiple": f"{_quote_sheet(ws.title)}!$B${start + 4}",
        "valuation.exit_revenue": f"{_quote_sheet(ws.title)}!$B${sensitivity_start + 1}",
    }


def _metric_reconciliation_rows(
    model: ModelBundle,
    lookup: dict[str, str],
) -> list[tuple[Any, ...]]:
    financials = model.financials
    transaction = model.transaction
    pipeline = model.assumptions["pipeline"]
    revenue_quality = model.assumptions["revenue_quality"]
    public_anchors = model.assumptions["public_anchors"]
    company_estimates = model.assumptions["company_estimates"]
    problem = model.assumptions["problem_benchmarks"]
    revenue_scale = financials.loc[2028, "revenue"] / financials.loc[2024, "revenue"]
    revenue_cagr = (
        financials.loc[2031, "revenue"] / financials.loc[2026, "revenue"]
    ) ** (1 / 5) - 1
    enterprise_runtime_mix = (
        financials.loc[2031, "revenue_enterprise"]
        + financials.loc[2031, "revenue_usage_and_deploy"]
    ) / financials.loc[2031, "revenue"]
    metrics = [
        (2, "Total raise", "$1.0B", transaction["total_raise"], "$M", "estimate", "MODEL-ESTIMATES", "transaction.total_raise"),
        (2, "Pre-money valuation", "$20.0B", transaction["pre_money"], "$M", "estimate", "MODEL-ESTIMATES", "transaction.pre_money"),
        (2, "Primary capital", "$800M", transaction["primary"], "$M", "estimate", "MODEL-ESTIMATES", "transaction.primary"),
        (2, "Secondary liquidity", "$200M", transaction["secondary"], "$M", "estimate", "MODEL-ESTIMATES", "transaction.secondary"),
        (2, "Post-money valuation", "$20.8B", transaction["post_money"], "$M", "derived", "MODEL-ESTIMATES", "transaction.post_money"),
        (2, "Primary dilution", "3.8%", transaction["primary_ownership"], "%", "derived", "MODEL-ESTIMATES", "transaction.primary_ownership"),
        (2, "Total buyer ownership", "4.8%", transaction["total_buyer_ownership"], "%", "derived", "MODEL-ESTIMATES", "transaction.total_buyer_ownership"),
        (2, "Pre-money / 2027E revenue", "16.1x", transaction["pre_money"] / financials.loc[2027, "revenue"], "x", "derived", "MODEL-ESTIMATES", "valuation.entry_multiple"),
        (4, "Registered users", "50M+", financials.loc[2026, "registered_users_m"], "M", "public", "RPL-D-2026", "kpi.registered_users_m.2026"),
        (4, "Fortune 500 presence", "85%", public_anchors["fortune_500_share"], "%", "public", "RPL-D-2026", "public_anchors.fortune_500_share"),
        (4, "Exit ARR target", "$1.0B", financials.loc[2026, "exit_arr"], "$M", "management_target", "RPL-D-2026", "kpi.exit_arr.2026"),
        (4, "2026 revenue", "$650M", financials.loc[2026, "revenue"], "$M", "estimate", "MODEL-ESTIMATES", "financial.revenue.2026"),
        (4, "Enterprise customers", "1,200", financials.loc[2026, "enterprise_customers"], "#", "estimate", "MODEL-ESTIMATES", "kpi.enterprise_customers.2026"),
        (4, "Countries reached", "190+", company_estimates["countries_reached"], "#", "estimate", "MODEL-ESTIMATES", "company_estimates.countries_reached"),
        (5, "Registered users", "50M+", financials.loc[2026, "registered_users_m"], "M", "public", "RPL-D-2026", "kpi.registered_users_m.2026"),
        (6, "Specialized tools", "6+", problem["specialized_tools"], "#", "estimate", "MODEL-ESTIMATES", "problem_benchmarks.specialized_tools"),
        (6, "Cross-functional handoffs", "7 to 12", problem["handoffs_high"], "#", "estimate", "MODEL-ESTIMATES", "problem_benchmarks.handoffs_high"),
        (6, "Idea-to-production cycle", "4 to 12 wks", problem["cycle_weeks_high"], "weeks", "estimate", "MODEL-ESTIMATES", "problem_benchmarks.cycle_weeks_high"),
        (6, "Agent test cost", "<$1", problem["agent_test_cost_usd"], "$", "estimate", "MODEL-ESTIMATES", "problem_benchmarks.agent_test_cost_usd"),
        (8, "Agent 4 speed improvement", "10x", public_anchors["agent_4_speed_multiple"], "x", "public", "RPL-AGENT4-2026", "public_anchors.agent_4_speed_multiple"),
        (9, "UKG feedback-capacity gain", "400%", public_anchors["ukg_feedback_capacity_multiple"], "x", "public", "RPL-CUST-UKG", "public_anchors.ukg_feedback_capacity_multiple"),
        (11, "TAM", "$250B", model.market["tam"].sum(), "$M", "derived", "MODEL-ESTIMATES", "market.total_tam"),
        (11, "Near-term SAM", "$90B", model.market["sam"].sum(), "$M", "derived", "MODEL-ESTIMATES", "market.total_sam"),
        (11, "2031 revenue / TAM", "2.1%", financials.loc[2031, "revenue"] / model.market["tam"].sum(), "%", "derived", "MODEL-ESTIMATES", None),
        (13, "2026 revenue", "$650M", financials.loc[2026, "revenue"], "$M", "estimate", "MODEL-ESTIMATES", "financial.revenue.2026"),
        (13, "2026 exit ARR", "$1.0B", financials.loc[2026, "exit_arr"], "$M", "management_target", "RPL-D-2026", "kpi.exit_arr.2026"),
        (13, "2024 to 2028 revenue growth", "73x", revenue_scale, "x", "derived", "MODEL-ESTIMATES", None),
        (13, "2028 registered users", "100M", financials.loc[2028, "registered_users_m"], "M", "estimate", "MODEL-ESTIMATES", "kpis.registered_users_m[5]"),
        (13, "2028 exit ARR", "$2.35B", financials.loc[2028, "exit_arr"], "$M", "estimate", "MODEL-ESTIMATES", "kpis.exit_arr[5]"),
        (14, "Recurring share", "70%", revenue_quality["recurring_share_2026"], "%", "estimate", "MODEL-ESTIMATES", "revenue_quality.recurring_share_2026"),
        (14, "Usage share", "30%", revenue_quality["usage_share_2026"], "%", "estimate", "MODEL-ESTIMATES", "revenue_quality.usage_share_2026"),
        (14, "Top-10 customer share", "3%", revenue_quality["top_10_customer_share_2026"], "%", "estimate", "MODEL-ESTIMATES", "revenue_quality.top_10_customer_share_2026"),
        (15, "Enterprise GRR", "92%", financials.loc[2026, "enterprise_grr"], "%", "estimate", "MODEL-ESTIMATES", "kpi.enterprise_grr.2026"),
        (15, "Enterprise NRR", "138%", financials.loc[2026, "enterprise_nrr"], "%", "estimate", "MODEL-ESTIMATES", "kpi.enterprise_nrr.2026"),
        (16, "UKG feedback-capacity gain", "400%", public_anchors["ukg_feedback_capacity_multiple"], "x", "public", "RPL-CUST-UKG", "public_anchors.ukg_feedback_capacity_multiple"),
        (16, "UKG employees", "16,000+", public_anchors["ukg_employees"], "#", "public", "RPL-CUST-UKG", "public_anchors.ukg_employees"),
        (18, "2026 gross margin", "46%", financials.loc[2026, "gross_margin"], "%", "estimate", "MODEL-ESTIMATES", "financial.gross_margin.2026"),
        (18, "Individual CAC payback", "6 mo.", model.unit_economics.loc[2026, "individual_payback_months"], "months", "estimate", "MODEL-ESTIMATES", "unit.individual_payback_months.2026"),
        (18, "Enterprise CAC payback", "14 mo.", model.unit_economics.loc[2026, "enterprise_payback_months"], "months", "estimate", "MODEL-ESTIMATES", "unit.enterprise_payback_months.2026"),
        (18, "Enterprise LTV / CAC", "4.2x", model.unit_economics.loc[2026, "enterprise_ltv_cac"], "x", "estimate", "MODEL-ESTIMATES", "unit.enterprise_ltv_cac.2026"),
        (20, "Qualified pipeline", "$880M", pipeline["qualified_pipeline"], "$M", "estimate", "MODEL-ESTIMATES", "pipeline.qualified_pipeline"),
        (20, "Late-stage pipeline", "$510M", pipeline["late_stage_pipeline"], "$M", "estimate", "MODEL-ESTIMATES", "pipeline.late_stage_pipeline"),
        (20, "Weighted pipeline", "$310M", pipeline["weighted_pipeline"], "$M", "estimate", "MODEL-ESTIMATES", "pipeline.weighted_pipeline"),
        (20, "New ARR target", "$260M", pipeline["next_12m_new_arr_target"], "$M", "estimate", "MODEL-ESTIMATES", "pipeline.next_12m_new_arr_target"),
        (20, "Pipeline coverage", "3.4x", pipeline["coverage"], "x", "estimate", "MODEL-ESTIMATES", "pipeline.coverage"),
        (20, "Win rate", "24%", pipeline["win_rate"], "%", "estimate", "MODEL-ESTIMATES", "pipeline.win_rate"),
        (20, "Median sales cycle", "78d", pipeline["median_sales_cycle_days"], "days", "estimate", "MODEL-ESTIMATES", "pipeline.median_sales_cycle_days"),
        (20, "Quota attainment", "64%", pipeline["quota_attainment"], "%", "estimate", "MODEL-ESTIMATES", "pipeline.quota_attainment"),
        (23, "2026 revenue", "$650M", financials.loc[2026, "revenue"], "$M", "estimate", "MODEL-ESTIMATES", "financial.revenue.2026"),
        (23, "2028 adjusted operating income", "$293M", financials.loc[2028, "adjusted_operating_income"], "$M", "estimate", "MODEL-ESTIMATES", "financial.adjusted_operating_income.2028"),
        (23, "2028 free cash flow", "$278M", financials.loc[2028, "free_cash_flow"], "$M", "estimate", "MODEL-ESTIMATES", "cash_flow.free_cash_flow.2028"),
        (23, "2031 adjusted operating margin", "31%", financials.loc[2031, "adjusted_operating_margin"], "%", "estimate", "MODEL-ESTIMATES", "financial.adjusted_operating_margin.2031"),
        (24, "2031 revenue", "$5.25B", financials.loc[2031, "revenue"], "$M", "estimate", "MODEL-ESTIMATES", "financial.revenue.2031"),
        (24, "2026 to 2031 revenue CAGR", "52%", revenue_cagr, "%", "derived", "MODEL-ESTIMATES", None),
        (24, "2031 gross margin", "76%", financials.loc[2031, "gross_margin"], "%", "estimate", "MODEL-ESTIMATES", "financial.gross_margin.2031"),
        (24, "2031 international mix", "52%", financials.loc[2031, "international_revenue_share"], "%", "estimate", "MODEL-ESTIMATES", "kpis.international_revenue_share[8]"),
        (24, "2031 enterprise + runtime mix", "77%", enterprise_runtime_mix, "%", "derived", "MODEL-ESTIMATES", None),
        (25, "2028 free cash flow", "$278M", financials.loc[2028, "free_cash_flow"], "$M", "estimate", "MODEL-ESTIMATES", "cash_flow.free_cash_flow.2028"),
        (26, "Primary allocation", "$800M", transaction["primary"], "$M", "estimate", "MODEL-ESTIMATES", "transaction.primary"),
        (27, "Cash at September 2026", "$274M", model.cash_summary["beginning_cash"], "$M", "estimate", "MODEL-ESTIMATES", "cash.beginning_cash"),
        (27, "Minimum operating cash", "$250M", model.cash_summary["minimum_cash"], "$M", "estimate", "MODEL-ESTIMATES", "cash.minimum_cash"),
        (27, "Cash falls below minimum without Series E", "Nov-26", model.cash_summary["funding_need_date"].to_pydatetime(), "date", "derived", "MODEL-ESTIMATES", "cash.funding_need_date"),
        (27, "Net primary proceeds", "$770M", model.cash_summary["net_primary_proceeds"], "$M", "derived", "MODEL-ESTIMATES", "cash.net_primary_proceeds"),
        (27, "Further equity before IPO", "None", model.cash_summary["next_equity_need"], "text", "derived", "MODEL-ESTIMATES", "cash.next_equity_need"),
        (27, "IPO target", "H2 2030", model.cash_summary["ipo_target_date"].to_pydatetime(), "date", "estimate", "MODEL-ESTIMATES", "cash.ipo_target_date"),
        (27, "Downside cash at IPO", "$902M", model.cash_summary["downside_cash_at_ipo"], "$M", "derived", "MODEL-ESTIMATES", "cash.downside_cash_at_ipo"),
        (30, "2031 revenue", "$5.25B", financials.loc[2031, "revenue"], "$M", "estimate", "MODEL-ESTIMATES", "financial.revenue.2031"),
    ]
    for _, use in model.use_of_proceeds.iterrows():
        key = str(use["category"]).lower().replace(" ", "_")
        metrics.append(
            (
                26,
                use["category"],
                f"${use['amount']:,.0f}M",
                use["amount"],
                "$M",
                "estimate",
                "MODEL-ESTIMATES",
                f"use.{key}",
            )
        )
    metrics = [
        (DISPLAY_SLIDE_NUMBER[item[0]], *item[1:])
        for item in metrics
        if item[0] in DISPLAY_SLIDE_NUMBER
    ]
    metrics.sort(key=lambda item: (item[0], item[1]))
    rows = []
    for slide, metric, display, value, units, evidence, source_ids, key in metrics:
        reference = lookup.get(key, "") if key else ""
        rows.append((slide, metric, display, value, units, evidence, source_ids, reference))
    return rows


def _reconciliation_sheet(
    ws: Any,
    model: ModelBundle,
    lookup: dict[str, str],
) -> None:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "Slide Reconciliation",
        "Headline metrics used in the presentation, with evidence and workbook-cell lineage.",
    )
    rows = _metric_reconciliation_rows(model, lookup)
    _write_table(
        ws,
        start,
        [
            "Slide",
            "Metric",
            "Deck display",
            "Model value",
            "Units",
            "Evidence class",
            "Source IDs",
            "Workbook reference",
            "Status",
        ],
        [(*row, "Linked" if row[-1] else "Source-only") for row in rows],
    )
    for row in range(start + 1, start + 1 + len(rows)):
        evidence = ws.cell(row, 6).value
        fill = {
            "public": BLUE_FILL,
            "derived": GRAY_FILL,
            "estimate": ESTIMATE_FILL,
            "management_target": GREEN_FILL,
        }.get(evidence, CREAM_FILL)
        for column in range(1, 10):
            ws.cell(row, column).fill = fill
        if ws.cell(row, 5).value == "%":
            ws.cell(row, 4).number_format = "0.0%"
        elif ws.cell(row, 5).value == "$M":
            ws.cell(row, 4).number_format = "$#,##0"
    _set_widths(ws, {1: 10, 2: 34, 3: 18, 4: 18, 5: 12, 6: 20, 7: 28, 8: 42, 9: 16})


def _qa_sheet(
    ws: Any,
    model: ModelBundle,
    sources: dict[str, Source],
) -> None:
    _prepare_sheet(ws, "A5")
    start = _sheet_title(
        ws,
        "QA Checks",
        "Build-time controls. PASS is required before the materials are considered final.",
    )
    tx = model.transaction
    checks = [
        (
            "Transaction sources = uses",
            tx["primary"] + tx["secondary"],
            tx["total_raise"],
            "PASS" if tx["primary"] + tx["secondary"] == tx["total_raise"] else "FAIL",
        ),
        (
            "Post-money valuation",
            tx["pre_money"] + tx["primary"],
            tx["post_money"],
            "PASS" if tx["pre_money"] + tx["primary"] == tx["post_money"] else "FAIL",
        ),
        (
            "Primary capital fully allocated",
            model.use_of_proceeds["amount"].sum(),
            tx["primary"],
            "PASS" if model.use_of_proceeds["amount"].sum() == tx["primary"] else "FAIL",
        ),
        (
            "TAM components sum",
            model.market["tam"].sum(),
            250000,
            "PASS" if model.market["tam"].sum() == 250000 else "FAIL",
        ),
        (
            "All public sources have URLs",
            sum(1 for source in sources.values() if source.evidence_class == "public" and source.url),
            sum(1 for source in sources.values() if source.evidence_class == "public"),
            "PASS"
            if all(source.url for source in sources.values() if source.evidence_class == "public")
            else "FAIL",
        ),
        (
            "Downside < base < upside revenue in 2031",
            model.scenarios["downside"].loc[2031, "revenue"],
            model.scenarios["upside"].loc[2031, "revenue"],
            "PASS"
            if model.scenarios["downside"].loc[2031, "revenue"]
            < model.scenarios["base"].loc[2031, "revenue"]
            < model.scenarios["upside"].loc[2031, "revenue"]
            else "FAIL",
        ),
        (
            "Base free cash flow positive by 2028",
            model.financials.loc[2028, "free_cash_flow"],
            0,
            "PASS" if model.financials.loc[2028, "free_cash_flow"] > 0 else "FAIL",
        ),
        (
            "Series E primary included once",
            model.monthly_cash_flow["primary_financing"].sum(),
            model.transaction["primary"],
            "PASS"
            if model.monthly_cash_flow["primary_financing"].sum()
            == model.transaction["primary"]
            else "FAIL",
        ),
        (
            "No-financing cash falls below minimum",
            model.cash_summary["funding_need_date"].to_pydatetime(),
            model.cash_summary["financing_close_date"].to_pydatetime(),
            "PASS" if model.cash_summary["funding_need_date"] is not None else "FAIL",
        ),
        (
            "Base cash remains above minimum",
            model.cash_summary["minimum_base_cash"],
            model.cash_summary["minimum_cash"],
            "PASS"
            if model.cash_summary["minimum_base_cash"]
            >= model.cash_summary["minimum_cash"]
            else "FAIL",
        ),
        (
            "Downside cash remains above minimum",
            model.cash_summary["minimum_downside_cash"],
            model.cash_summary["minimum_cash"],
            "PASS"
            if model.cash_summary["minimum_downside_cash"]
            >= model.cash_summary["minimum_cash"]
            else "FAIL",
        ),
    ]
    _write_table(
        ws,
        start,
        ["Check", "Observed", "Expected / comparator", "Status"],
        checks,
    )
    for row in range(start + 1, start + 1 + len(checks)):
        status_cell = ws.cell(row, 4)
        fill = GREEN_FILL if status_cell.value == "PASS" else RED_FILL
        for column in range(1, 5):
            ws.cell(row, column).fill = fill
    ws.conditional_formatting.add(
        f"D{start + 1}:D{start + len(checks)}",
        CellIsRule(operator="equal", formula=['"FAIL"'], fill=RED_FILL),
    )
    _set_widths(ws, {1: 46, 2: 22, 3: 24, 4: 14})


def build_workbook(
    model: ModelBundle,
    sources: dict[str, Source],
    output_path: Path = XLSX_PATH,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in SHEET_NAMES:
        workbook.create_sheet(sheet_name)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    lookup: dict[str, str] = {}
    lookup.update(_assumptions_sheet(workbook["Assumptions"], model))
    _visuals_sheet(workbook["Visuals"], model)
    lookup.update(_p_and_l_sheet(workbook["P&L"], model))
    lookup.update(_cash_flow_sheet(workbook["Cash Flow"], model))
    lookup.update(_balance_sheet(workbook["Balance Sheet"], model))
    lookup.update(_read_me_sheet(workbook["Read Me"], model))
    _source_register_sheet(workbook["Source Register"], sources)
    _public_facts_sheet(workbook["Public Facts"], sources)
    lookup.update(_historical_kpis_sheet(workbook["Historical KPIs"], model))
    lookup.update(_revenue_build_sheet(workbook["Revenue Build"], model))
    _cohorts_sheet(workbook["Cohorts & Retention"], model)
    lookup.update(_unit_economics_sheet(workbook["Unit Economics"], model))
    lookup.update(_pipeline_sheet(workbook["GTM & Pipeline"], model))
    lookup.update(_market_sheet(workbook["Market Model"], model))
    lookup.update(_scenarios_sheet(workbook["Scenarios"], model))
    lookup.update(_use_of_proceeds_sheet(workbook["Use of Proceeds"], model))
    lookup.update(_valuation_sheet(workbook["Valuation"], model))
    _reconciliation_sheet(workbook["Slide Reconciliation"], model, lookup)
    _qa_sheet(workbook["QA Checks"], model, sources)

    workbook.properties.title = "Replit Series E Sourcebook"
    workbook.properties.subject = "Illustrative investor underwriting model"
    workbook.properties.creator = "Replit Series E materials generator"
    workbook.save(output_path)
    return output_path
